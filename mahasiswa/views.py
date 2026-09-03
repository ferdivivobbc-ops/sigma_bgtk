from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.db import IntegrityError
from django.views.decorators.http import require_POST
from django.utils import timezone
from datetime import timedelta
from accounts.models import User
from accounts.decorators import admin_required, role_required
from kampus.models import Kampus
from mahasiswa.models import Mahasiswa
from magang.models import Magang
from django.db import transaction
from openpyxl import load_workbook


def _unique_username(email, nim):
    base_username = email.split('@')[0] + '_' + nim[-4:]
    username = base_username
    suffix = 2
    while User.objects.filter(username=username).exists():
        username = f'{base_username}_{suffix}'
        suffix += 1
    return username

@login_required
@role_required('ADMIN', 'PEMBIMBING')
def mahasiswa_list(request):
    search_query = request.GET.get('q', '').strip()
    kampus_filter = request.GET.get('kampus', '').strip()
    status_filter = request.GET.get('status', '').strip()

    mahasiswa_qs = Mahasiswa.objects.select_related('user', 'kampus', 'data_magang').all()

    if request.user.is_pembimbing():
        # Filter only students assigned to this supervisor
        mahasiswa_qs = mahasiswa_qs.filter(data_magang__pembimbing=request.user)

    if search_query:
        mahasiswa_qs = mahasiswa_qs.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(nim__icontains=search_query) |
            Q(prodi__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )

    if kampus_filter:
        mahasiswa_qs = mahasiswa_qs.filter(kampus_id=kampus_filter)

    if status_filter:
        mahasiswa_qs = mahasiswa_qs.filter(status=status_filter)

    paginator = Paginator(mahasiswa_qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    kampus_list = Kampus.objects.filter(status='AKTIF')

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'kampus_filter': kampus_filter,
        'status_filter': status_filter,
        'kampus_list': kampus_list,
    }
    return render(request, 'mahasiswa/mahasiswa_list.html', context)

@login_required
@role_required('ADMIN', 'PEMBIMBING')
def mahasiswa_detail(request, pk):
    mhs = get_object_or_404(Mahasiswa.objects.select_related('user', 'kampus', 'data_magang__pembimbing'), pk=pk)
    magang = getattr(mhs, 'data_magang', None)
    absensi_list = mhs.absensi_list.all()[:10]
    kegiatan_list = mhs.kegiatan_harian_list.all()[:10]

    context = {
        'mahasiswa': mhs,
        'magang': magang,
        'absensi_list': absensi_list,
        'kegiatan_list': kegiatan_list,
    }
    return render(request, 'mahasiswa/mahasiswa_detail.html', context)

@login_required
@role_required('ADMIN', 'PEMBIMBING')
def mahasiswa_create(request):
    if request.user.is_pembimbing():
        search_query = request.GET.get('q', '').strip()
        if request.method == 'POST':
            mahasiswa_id = request.POST.get('mahasiswa_id', '')
            mhs = get_object_or_404(Mahasiswa, pk=mahasiswa_id)
            existing_magang = getattr(mhs, 'data_magang', None)
            if existing_magang and existing_magang.pembimbing_id not in (None, request.user.pk):
                messages.error(request, 'Mahasiswa tersebut sudah terhubung ke pembimbing lain.')
            else:
                start_date = timezone.localdate()
                Magang.objects.update_or_create(
                    mahasiswa=mhs,
                    defaults={
                        'tanggal_mulai': start_date,
                        'tanggal_selesai': start_date + timedelta(days=90),
                        'bagian': existing_magang.bagian if existing_magang else 'Magang',
                        'pembimbing': request.user,
                        'status': 'AKTIF',
                    }
                )
                messages.success(request, f'{mhs.user.get_full_name()} berhasil ditambahkan ke mahasiswa bimbingan.')
                return redirect('mahasiswa_list')

        mahasiswa_search = Mahasiswa.objects.select_related('user', 'kampus').all()
        if search_query:
            search_filter = Q()
            for search_term in search_query.split():
                search_filter &= (
                    Q(user__first_name__icontains=search_term) |
                    Q(user__last_name__icontains=search_term) |
                    Q(nim__icontains=search_term) |
                    Q(kampus__nama_kampus__icontains=search_term)
                )
            mahasiswa_search = mahasiswa_search.filter(search_filter)
        mahasiswa_search = mahasiswa_search.order_by('user__first_name', 'nim')[:30]
        return render(request, 'mahasiswa/mahasiswa_assign.html', {
            'mahasiswa_search': mahasiswa_search,
            'search_query': search_query,
        })

    if request.method == 'POST':
        full_name = request.POST.get('full_name', '').strip()
        email = request.POST.get('email', '').strip()
        nim = request.POST.get('nim', '').strip()
        prodi = request.POST.get('prodi', '').strip()
        kampus_id = request.POST.get('kampus', '')
        no_hp = request.POST.get('no_hp', '').strip()
        password = request.POST.get('password', '').strip() or 'mahasiswa123'
        status = request.POST.get('status', 'AKTIF')
        pembimbing_id = request.POST.get('pembimbing', '')
        tanggal_mulai = request.POST.get('tanggal_mulai', '')
        tanggal_selesai = request.POST.get('tanggal_selesai', '')
        bagian = request.POST.get('bagian', '').strip()
        if request.user.is_pembimbing():
            pembimbing_id = request.user.pk

        if not email or not nim or not full_name or not kampus_id:
            messages.error(request, "Nama, Email, NIM, dan Kampus wajib diisi.")
        elif Mahasiswa.objects.filter(nim=nim).exists():
            messages.error(request, "NIM tersebut sudah terdaftar.")
        elif request.user.is_pembimbing() and (not tanggal_mulai or not tanggal_selesai or not bagian):
            messages.error(request, "Tanggal mulai, tanggal selesai, dan bagian wajib diisi jika pembimbing dipilih.")
        else:
            try:
                with transaction.atomic():
                    user = User.objects.create_user(
                        username=_unique_username(email, nim),
                        email=email,
                        password=password,
                        first_name=full_name,
                        role=User.Role.MAHASISWA,
                        no_hp=no_hp
                    )

                    mhs = Mahasiswa.objects.create(
                        user=user,
                        kampus_id=kampus_id,
                        nim=nim,
                        prodi=prodi,
                        no_hp=no_hp,
                        status=status
                    )
                    if 'foto' in request.FILES:
                        mhs.foto = request.FILES['foto']
                        mhs.save()
                    if pembimbing_id:
                        Magang.objects.create(
                            mahasiswa=mhs,
                            tanggal_mulai=tanggal_mulai,
                            tanggal_selesai=tanggal_selesai,
                            bagian=bagian,
                            pembimbing_id=pembimbing_id,
                            status='AKTIF',
                        )
            except IntegrityError:
                messages.error(request, "Data mahasiswa gagal dibuat karena email, NIM, atau username sudah terdaftar.")
            else:
                messages.success(request, f"Mahasiswa {full_name} berhasil dibuat.")
                return redirect('mahasiswa_list')

    kampus_list = Kampus.objects.filter(status='AKTIF')
    pembimbing_list = User.objects.filter(role=User.Role.PEMBIMBING)
    return render(request, 'mahasiswa/mahasiswa_form.html', {
        'kampus_list': kampus_list,
        'pembimbing_list': pembimbing_list,
        'action': 'Tambah',
    })

@login_required
@admin_required
def mahasiswa_edit(request, pk):
    mhs = get_object_or_404(Mahasiswa.objects.select_related('user'), pk=pk)

    if request.method == 'POST':
        full_name = request.POST.get('full_name', '').strip()
        email = request.POST.get('email', '').strip()
        nim = request.POST.get('nim', '').strip()
        prodi = request.POST.get('prodi', '').strip()
        kampus_id = request.POST.get('kampus', '')
        no_hp = request.POST.get('no_hp', '').strip()
        status = request.POST.get('status', 'AKTIF')
        pembimbing_id = request.POST.get('pembimbing', '')
        tanggal_mulai = request.POST.get('tanggal_mulai', '')
        tanggal_selesai = request.POST.get('tanggal_selesai', '')
        bagian = request.POST.get('bagian', '').strip()

        mhs.user.first_name = full_name
        mhs.user.email = email
        mhs.user.no_hp = no_hp
        mhs.user.save()

        mhs.nim = nim
        mhs.prodi = prodi
        mhs.kampus_id = kampus_id
        mhs.no_hp = no_hp
        mhs.status = status
        if 'foto' in request.FILES:
            mhs.foto = request.FILES['foto']
        mhs.save()

        if pembimbing_id:
            if not tanggal_mulai or not tanggal_selesai or not bagian:
                messages.error(request, "Tanggal mulai, tanggal selesai, dan bagian wajib diisi jika pembimbing dipilih.")
                return redirect('mahasiswa_edit', pk=mhs.pk)
            Magang.objects.update_or_create(
                mahasiswa=mhs,
                defaults={
                    'tanggal_mulai': tanggal_mulai,
                    'tanggal_selesai': tanggal_selesai,
                    'bagian': bagian,
                    'pembimbing_id': pembimbing_id,
                    'status': 'AKTIF',
                }
            )

        messages.success(request, f"Data mahasiswa {full_name} berhasil diperbarui.")
        return redirect('mahasiswa_list')

    kampus_list = Kampus.objects.filter(status='AKTIF')
    pembimbing_list = User.objects.filter(role=User.Role.PEMBIMBING)
    return render(request, 'mahasiswa/mahasiswa_form.html', {
        'mahasiswa': mhs,
        'kampus_list': kampus_list,
        'pembimbing_list': pembimbing_list,
        'action': 'Edit',
    })


@login_required
@admin_required
@require_POST
def mahasiswa_delete(request, pk):
    mhs = get_object_or_404(Mahasiswa.objects.select_related('user'), pk=pk)
    name = mhs.user.get_full_name() or mhs.user.username
    mhs.delete()
    messages.success(request, f"Data mahasiswa {name} berhasil dihapus.")
    return redirect('mahasiswa_list')


@login_required
@admin_required
def mahasiswa_import(request):
    if request.method == 'POST':
        uploaded_file = request.FILES.get('file_excel')
        if not uploaded_file or not uploaded_file.name.lower().endswith('.xlsx'):
            messages.error(request, 'Pilih file Excel .xlsx terlebih dahulu.')
            return render(request, 'mahasiswa/mahasiswa_import.html')
        if uploaded_file.size > 10 * 1024 * 1024:
            messages.error(request, 'Ukuran file maksimal 10 MB.')
            return render(request, 'mahasiswa/mahasiswa_import.html')

        try:
            workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            headers = [str(value or '').strip().lower() for value in (rows[0] if rows else [])]
            required_headers = {'nama', 'email', 'nim', 'prodi', 'kampus'}
            if not required_headers.issubset(set(headers)):
                messages.error(request, 'Header wajib: nama, email, nim, prodi, kampus. Kolom no_hp opsional.')
                return render(request, 'mahasiswa/mahasiswa_import.html')

            positions = {header: headers.index(header) for header in headers}
            imported_count = 0
            errors = []
            with transaction.atomic():
                for row_number, row in enumerate(rows[1:], start=2):
                    values = {header: row[index] if index < len(row) else '' for header, index in positions.items()}
                    name = str(values.get('nama') or '').strip()
                    email = str(values.get('email') or '').strip().lower()
                    nim = str(values.get('nim') or '').strip()
                    prodi = str(values.get('prodi') or '').strip()
                    campus_name = str(values.get('kampus') or '').strip()
                    no_hp = str(values.get('no_hp') or '').strip()
                    if not all([name, email, nim, prodi, campus_name]):
                        errors.append(f'Baris {row_number}: data wajib belum lengkap.')
                        continue
                    if User.objects.filter(email=email).exists() or Mahasiswa.objects.filter(nim=nim).exists():
                        errors.append(f'Baris {row_number}: email atau NIM sudah terdaftar.')
                        continue
                    kampus = Kampus.objects.filter(nama_kampus__iexact=campus_name, status='AKTIF').first()
                    if not kampus:
                        errors.append(f'Baris {row_number}: kampus "{campus_name}" tidak ditemukan.')
                        continue
                    username = email.split('@')[0] + '_' + nim[-4:]
                    user = User.objects.create_user(username=username, email=email, password='mahasiswa123', first_name=name, role=User.Role.MAHASISWA, no_hp=no_hp)
                    Mahasiswa.objects.create(user=user, kampus=kampus, nim=nim, prodi=prodi, no_hp=no_hp, status='AKTIF')
                    imported_count += 1
            workbook.close()
        except Exception:
            messages.error(request, 'File Excel tidak dapat dibaca. Pastikan formatnya .xlsx.')
            return render(request, 'mahasiswa/mahasiswa_import.html')

        if imported_count:
            messages.success(request, f'{imported_count} mahasiswa berhasil diimport. Password awal: mahasiswa123.')
        if errors:
            messages.warning(request, 'Beberapa baris dilewati: ' + ' '.join(errors[:5]))
        return redirect('mahasiswa_list')

    return render(request, 'mahasiswa/mahasiswa_import.html')
