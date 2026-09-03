"""
Utility untuk export data ke format Excel
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


class ExcelExporter:
    """Class untuk membuat file Excel dengan formatting"""
    
    def __init__(self, filename="export"):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.filename = filename
        self.row = 1
        
    def set_title(self, title):
        """Set judul worksheet"""
        self.ws.title = title[:31]  # Max 31 karakter untuk sheet name
        
    def add_header(self, headers, col_widths=None):
        """Tambah header row dengan styling"""
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=self.row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            
        if col_widths:
            for col_num, width in enumerate(col_widths, 1):
                self.ws.column_dimensions[get_column_letter(col_num)].width = width
                
        self.row += 1
        
    def add_row(self, values, is_bold=False):
        """Tambah data row"""
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, value in enumerate(values, 1):
            cell = self.ws.cell(row=self.row, column=col_num)
            cell.value = value
            cell.border = border
            if is_bold:
                cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
        self.row += 1
        
    def add_rows(self, data_list):
        """Tambah multiple rows dari list"""
        for data in data_list:
            self.add_row(data)
            
    def freeze_panes(self, row=2, col=1):
        """Freeze panes untuk header"""
        self.ws.freeze_panes = f"{get_column_letter(col)}{row}"
        
    def get_response(self):
        """Generate HTTP response dengan file Excel"""
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        response['Content-Disposition'] = f'attachment; filename="{self.filename}_{timestamp}.xlsx"'
        return response


def export_mahasiswa_to_excel(queryset):
    """Export mahasiswa data ke Excel"""
    exporter = ExcelExporter("daftar_mahasiswa")
    exporter.set_title("Mahasiswa")
    
    headers = ["No", "NIM", "Nama", "Email", "No HP", "Kampus", "Prodi", "Tanggal Masuk"]
    col_widths = [5, 15, 25, 25, 15, 25, 20, 15]
    exporter.add_header(headers, col_widths)
    
    for idx, mhs in enumerate(queryset, 1):
        exporter.add_row([
            idx,
            mhs.nim,
            mhs.user.get_full_name(),
            mhs.user.email,
            mhs.no_hp or "-",
            mhs.kampus.nama_kampus if mhs.kampus else "-",
            mhs.prodi or "-",
            mhs.created_at.strftime("%d-%m-%Y") if mhs.created_at else "-"
        ])
    
    exporter.freeze_panes()
    return exporter.get_response()


def export_absensi_to_excel(queryset):
    """Export absensi data ke Excel"""
    exporter = ExcelExporter("daftar_absensi")
    exporter.set_title("Absensi")
    
    headers = ["No", "Tanggal", "Nama Mahasiswa", "NIM", "Status", "Jam Masuk", "Jam Pulang", "Keterangan"]
    col_widths = [5, 15, 25, 15, 15, 15, 15, 20]
    exporter.add_header(headers, col_widths)
    
    for idx, absensi in enumerate(queryset.order_by('-tanggal'), 1):
        exporter.add_row([
            idx,
            absensi.tanggal.strftime("%d-%m-%Y"),
            absensi.mahasiswa.user.get_full_name(),
            absensi.mahasiswa.nim,
            absensi.status,
            absensi.jam_masuk.strftime("%H:%M") if absensi.jam_masuk else "-",
            absensi.jam_pulang.strftime("%H:%M") if absensi.jam_pulang else "-",
            absensi.keterangan or "-"
        ])
    
    exporter.freeze_panes()
    return exporter.get_response()


def export_kegiatan_to_excel(queryset):
    """Export kegiatan harian data ke Excel"""
    exporter = ExcelExporter("daftar_kegiatan")
    exporter.set_title("Kegiatan Harian")
    
    headers = ["No", "Tanggal", "Nama Mahasiswa", "Judul Kegiatan", "Uraian", "Hasil", "Status", "Verifikasi Oleh"]
    col_widths = [5, 15, 25, 20, 25, 20, 15, 20]
    exporter.add_header(headers, col_widths)
    
    for idx, kegiatan in enumerate(queryset.order_by('-tanggal'), 1):
        exporter.add_row([
            idx,
            kegiatan.tanggal.strftime("%d-%m-%Y"),
            kegiatan.mahasiswa.user.get_full_name(),
            kegiatan.judul_kegiatan,
            kegiatan.kegiatan[:50] + "..." if len(kegiatan.kegiatan) > 50 else kegiatan.kegiatan,
            kegiatan.hasil[:50] + "..." if len(kegiatan.hasil) > 50 else kegiatan.hasil,
            kegiatan.status_verifikasi,
            kegiatan.verified_by.get_full_name() if kegiatan.verified_by else "-"
        ])
    
    exporter.freeze_panes()
    return exporter.get_response()


def export_magang_to_excel(queryset):
    """Export data magang ke Excel"""
    exporter = ExcelExporter("daftar_magang")
    exporter.set_title("Magang")
    
    headers = ["No", "Nama Mahasiswa", "NIM", "Bagian", "Pembimbing", "Tanggal Mulai", "Tanggal Selesai", "Status"]
    col_widths = [5, 25, 15, 20, 20, 15, 15, 15]
    exporter.add_header(headers, col_widths)
    
    for idx, magang in enumerate(queryset, 1):
        exporter.add_row([
            idx,
            magang.mahasiswa.user.get_full_name(),
            magang.mahasiswa.nim,
            magang.bagian or "-",
            magang.pembimbing.get_full_name() if magang.pembimbing else "-",
            magang.tanggal_mulai.strftime("%d-%m-%Y") if magang.tanggal_mulai else "-",
            magang.tanggal_selesai.strftime("%d-%m-%Y") if magang.tanggal_selesai else "-",
            magang.status or "-"
        ])
    
    exporter.freeze_panes()
    return exporter.get_response()
