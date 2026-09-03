"""
Views untuk export data ke berbagai format (Excel, PDF, etc)
"""
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from datetime import datetime, timedelta

from accounts.decorators import role_required, pembimbing_required
from mahasiswa.models import Mahasiswa
from absensi.models import Absensi
from kegiatan.models import KegiatanHarian
from magang.models import Magang

from utils.export_excel import (
    export_mahasiswa_to_excel,
    export_absensi_to_excel,
    export_kegiatan_to_excel,
    export_magang_to_excel
)


@login_required
@role_required('ADMIN', 'PEMBIMBING')
def export_mahasiswa_excel(request):
    """Export daftar mahasiswa ke Excel"""
    # Filter by pembimbing jika user adalah pembimbing
    if request.user.is_pembimbing():
        mahasiswa_qs = Mahasiswa.objects.filter(
            data_magang__pembimbing=request.user
        ).distinct()
    else:
        mahasiswa_qs = Mahasiswa.objects.all()
    
    # Filter by kampus jika ada parameter
    kampus_id = request.GET.get('kampus')
    if kampus_id:
        mahasiswa_qs = mahasiswa_qs.filter(kampus_id=kampus_id)
    
    return export_mahasiswa_to_excel(mahasiswa_qs)


@login_required
@role_required('ADMIN', 'PEMBIMBING')
def export_absensi_excel(request):
    """Export data absensi ke Excel"""
    absensi_qs = Absensi.objects.select_related('mahasiswa__user', 'mahasiswa__kampus')
    
    # Filter by pembimbing
    if request.user.is_pembimbing():
        absensi_qs = absensi_qs.filter(
            mahasiswa__data_magang__pembimbing=request.user
        )
    
    # Filter by date range
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if from_date:
        absensi_qs = absensi_qs.filter(tanggal__gte=from_date)
    if to_date:
        absensi_qs = absensi_qs.filter(tanggal__lte=to_date)
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        absensi_qs = absensi_qs.filter(status=status)
    
    return export_absensi_to_excel(absensi_qs)


@login_required
@role_required('ADMIN', 'PEMBIMBING')
def export_kegiatan_excel(request):
    """Export data kegiatan harian ke Excel"""
    kegiatan_qs = KegiatanHarian.objects.select_related(
        'mahasiswa__user',
        'mahasiswa__kampus',
        'verified_by'
    )
    
    # Filter by pembimbing
    if request.user.is_pembimbing():
        kegiatan_qs = kegiatan_qs.filter(
            mahasiswa__data_magang__pembimbing=request.user
        )
    
    # Filter by date range
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if from_date:
        kegiatan_qs = kegiatan_qs.filter(tanggal__gte=from_date)
    if to_date:
        kegiatan_qs = kegiatan_qs.filter(tanggal__lte=to_date)
    
    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter:
        kegiatan_qs = kegiatan_qs.filter(status_verifikasi=status_filter)
    
    return export_kegiatan_to_excel(kegiatan_qs)


@login_required
@role_required('ADMIN', 'PEMBIMBING')
def export_magang_excel(request):
    """Export data magang ke Excel"""
    magang_qs = Magang.objects.select_related(
        'mahasiswa__user',
        'mahasiswa__kampus',
        'pembimbing'
    )
    
    # Filter by pembimbing
    if request.user.is_pembimbing():
        magang_qs = magang_qs.filter(pembimbing=request.user)
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        magang_qs = magang_qs.filter(status=status)
    
    return export_magang_to_excel(magang_qs)


@login_required
def export_rekap_kegiatan_bulanan_excel(request):
    """Export rekap kegiatan bulanan per mahasiswa ke Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from kegiatan.models import KegiatanBulanan
    
    # Jika user adalah mahasiswa, hanya bisa export data sendiri
    if request.user.is_mahasiswa():
        mahasiswa = request.user.mahasiswa_profile
        rekap_qs = KegiatanBulanan.objects.filter(mahasiswa=mahasiswa)
    else:
        # Admin/Pembimbing bisa filter
        if request.user.is_pembimbing():
            rekap_qs = KegiatanBulanan.objects.filter(
                mahasiswa__data_magang__pembimbing=request.user
            )
        else:
            rekap_qs = KegiatanBulanan.objects.all()
    
    bulan = request.GET.get('bulan')
    tahun = request.GET.get('tahun')
    
    if bulan:
        rekap_qs = rekap_qs.filter(bulan=int(bulan))
    if tahun:
        rekap_qs = rekap_qs.filter(tahun=int(tahun))
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Kegiatan Bulanan"
    
    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["No", "Nama Mahasiswa", "Bulan", "Tahun", "Total Hadir", "Total Kegiatan", "Total Disetujui", "Ringkasan"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 40
    
    # Data rows
    for idx, rekap in enumerate(rekap_qs, 1):
        row = idx + 1
        ws.cell(row=row, column=1).value = idx
        ws.cell(row=row, column=2).value = rekap.mahasiswa.user.get_full_name()
        ws.cell(row=row, column=3).value = rekap.bulan
        ws.cell(row=row, column=4).value = rekap.tahun
        ws.cell(row=row, column=5).value = rekap.total_hadir
        ws.cell(row=row, column=6).value = rekap.total_kegiatan
        ws.cell(row=row, column=7).value = rekap.total_disetujui
        ws.cell(row=row, column=8).value = rekap.ringkasan[:100] if rekap.ringkasan else "-"
        
        # Apply border to all cells
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = border
    
    # Freeze panes
    ws.freeze_panes = "A2"
    
    # Generate response
    output = __import__('io').BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    response['Content-Disposition'] = f'attachment; filename="rekap_kegiatan_bulanan_{timestamp}.xlsx"'
    
    return response
